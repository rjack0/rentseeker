#!/usr/bin/env python3
"""Convert LA County SBF workbook parts to CSV files for DuckDB.

The Electron owner service reads sbf_part1.csv, sbf_part2.csv, and
sbf_part3.csv from the SBF source directory. This script performs that
one-time conversion without loading each workbook fully into memory.
"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook


SBF_DIR = Path(
    "/Users/rjack/Desktop/almanac/Docs/RE Data/"
    "SBF Secured Basic File LA County Assessor Abstract"
)

WORKBOOKS = [
    ("Custom DS04 Part 1 Data Sales.xlsx", "sbf_part1.csv"),
    ("Custom DS04 Part 2 Data Sales.xlsx", "sbf_part2.csv"),
    ("Custom DS04 Part 3 Data Sales.xlsx", "sbf_part3.csv"),
]


def convert_workbook(source: Path, destination: Path) -> int:
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    row_count = 0

    with destination.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(["" if value is None else value for value in row])
            row_count += 1

    workbook.close()
    return max(0, row_count - 1)


def main() -> None:
    for source_name, destination_name in WORKBOOKS:
        source = SBF_DIR / source_name
        destination = SBF_DIR / destination_name
        if not source.exists():
            raise FileNotFoundError(source)
        rows = convert_workbook(source, destination)
        print(f"{destination}: {rows:,} data rows")


if __name__ == "__main__":
    main()
